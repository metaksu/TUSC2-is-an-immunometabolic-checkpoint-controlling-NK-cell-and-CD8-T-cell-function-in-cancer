"""Merge the collected Figure 7 Supplementary data/table files into ONE .xlsx,
one sheet per file, preceded by a Contents index sheet. Reads the tumor-only
collected files in supplementary/ (so it matches the shipped set exactly);
re-run after collect_supplementary.py.

Run: python make_supplementary_xlsx.py
Out: supplementary/Figure7_Supplementary_Data.xlsx
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
SUPP = HERE / "supplementary"
OUT = SUPP / "Figure7_Supplementary_Data.xlsx"

# (sheet name <=31 chars, Suppl. label, description, filename, kind)
SHEETS = [
    ("S1_NK_per_gene_DE",        "Suppl. Data S1",  "NK per-gene DESeq2 pseudobulk DE (Tang_Tumor)",       "SupplData_S1_NK_per_gene_DE.csv",              "csv"),
    ("S1_CD8_per_gene_DE",       "Suppl. Data S1",  "CD8 per-gene DESeq2 pseudobulk DE (CD8_TIL)",         "SupplData_S1_CD8_per_gene_DE.csv",             "csv"),
    ("S2_NK_effect_size_ladder", "Suppl. Data S2",  "NK effect-size ladder + calibrated 1.1-fold null",    "SupplData_S2_NK_effect_size_ladder.tsv",       "tsv"),
    ("S2_CD8_effect_size_ladder","Suppl. Data S2",  "CD8 effect-size ladder + calibrated 1.1-fold null",   "SupplData_S2_CD8_effect_size_ladder.tsv",      "tsv"),
    ("S3_NK_GSEA",               "Suppl. Data S3",  "NK GSEA preranked (res2d)",                           "SupplData_S3_NK_GSEA.csv",                     "csv"),
    ("S3_CD8_GSEA",              "Suppl. Data S3",  "CD8 GSEA preranked (res2d)",                          "SupplData_S3_CD8_GSEA.csv",                    "csv"),
    ("S4_NK_label_provenance",   "Suppl. Data S4",  "NK curated volcano-label provenance (Tang_Tumor)",    "SupplData_S4_NK_curated_label_provenance.tsv", "tsv"),
    ("S4_CD8_label_provenance",  "Suppl. Data S4",  "CD8 curated volcano-label provenance (CD8_TIL)",      "SupplData_S4_CD8_curated_label_provenance.tsv","tsv"),
    ("Table_S1_availability",    "Suppl. Table S1", "Data and code availability",                          "SupplTable_S1_data_code_availability.md",      "md"),
]


def read_md_table(path: Path) -> pd.DataFrame:
    """Parse the single markdown pipe-table in the availability .md (strip bold)."""
    tbl = [l for l in path.read_text(encoding="utf-8").splitlines() if l.strip().startswith("|")]
    def cells(row):
        return [c.strip().replace("**", "") for c in row.strip().strip("|").split("|")]
    header = cells(tbl[0])
    rows = [cells(r) for r in tbl[2:]]        # skip the |---| separator row
    return pd.DataFrame(rows, columns=header)


def load(kind: str, path: Path) -> pd.DataFrame:
    if kind == "csv":
        return pd.read_csv(path)
    if kind == "tsv":
        return pd.read_csv(path, sep="\t")
    return read_md_table(path)


def main() -> int:
    frames, contents, missing = [], [], []
    for sheet, label, desc, fname, kind in SHEETS:
        path = SUPP / fname
        if not path.exists():
            missing.append(fname); print(f"  [MISSING] {fname}"); continue
        df = load(kind, path)
        frames.append((sheet, df))
        contents.append({"Sheet": sheet, "Suppl. label": label, "Description": desc,
                         "Source file": fname, "Rows": len(df), "Columns": df.shape[1]})
        print(f"  [ok] {sheet}: {len(df)} rows x {df.shape[1]} cols  <- {fname}")

    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        pd.DataFrame(contents).to_excel(xw, sheet_name="Contents", index=False)
        for sheet, df in frames:
            df.to_excel(xw, sheet_name=sheet, index=False)
        wb = xw.book
        for ws in wb.worksheets:                       # header polish + usability
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            if ws.max_column >= 1 and ws.max_row >= 1:
                ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:                     # rough auto-width (sample first 200 rows)
                w = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(60, max(10, w + 2))

    print(f"\nwrote {OUT}")
    print(f"  {len(frames)} data sheets + Contents = {len(frames) + 1} tabs")
    if missing:
        print(f"  MISSING ({len(missing)}): {missing}")
    return 1 if missing else 0


if __name__ == "__main__":
    raise SystemExit(main())
